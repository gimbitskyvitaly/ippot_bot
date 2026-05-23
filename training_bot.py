#!/usr/bin/env python3
"""
Telegram бот для опросов о тренировках.
Бот отправляет опросы за 4 дня до тренировки (пятница и воскресенье)
и сохраняет результаты в Excel таблицу.

После тренировки (через 2 часа) бот рассылает запросы на оценку
тренировки (1-10) и сбор фидбэка участникам.
Статистика оценок и текстовые отзывы сохраняются в feedback.xlsx.
"""

import logging
import asyncio
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook, load_workbook
from telegram import Update
from telegram.ext import Application, CommandHandler, PollAnswerHandler, ContextTypes, MessageHandler, filters

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# КОНФИГУРАЦИЯ
# Замените на ваш токен бота от @BotFather
BOT_TOKEN = "YOUR_BOT_TOKEN_HERE"
# ID группы, куда добавлять бота (можно найти через @userinfobot или добавив бота в группу)
GROUP_ID = "YOUR_GROUP_ID_HERE"
# ID ветки (forum topic) для опросов в группе
THREAD_ID = None  # Укажите ID темы "опросы", если группа имеет формат форума
# Путь к файлу Excel для хранения статистики
EXCEL_FILE = "attendance.xlsx"


# Глобальное хранилище для связи ID опроса с данными тренировки
# Формат: { poll_id: {"date": "DD.MM", "time": "HH:MM", "location": "LOC"} }
active_polls = {}

# Хранилище для пользователей, записавшихся на тренировку
# Формат: { "DD.MM_HH:MM": {user_id: username} }
training_attendees = {}

# Хранилище для отправленных запросов на фидбэк (чтобы не дублировать)
# Формат: { "DD.MM_HH:MM": True }
feedback_requests_sent = set()


def get_next_friday():
    """Возвращает дату следующей пятницы."""
    today = datetime.now().date()
    days_until_friday = (4 - today.weekday()) % 7
    if days_until_friday == 0:
        days_until_friday = 7
    return today + timedelta(days=days_until_friday)


def get_next_sunday():
    """Возвращает дату следующего воскресенья."""
    today = datetime.now().date()
    days_until_sunday = (6 - today.weekday()) % 7
    if days_until_sunday == 0:
        days_until_sunday = 7
    return today + timedelta(days=days_until_sunday)


def init_excel_file():
    """Инициализирует Excel файл с заголовками."""
    excel_path = Path(EXCEL_FILE)
    
    if not excel_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Посещаемость"
        
        # Заголовки
        ws.cell(row=1, column=1, value="Имя")
        ws.cell(row=1, column=2, value="Посещаемость (%)")
        
        wb.save(EXCEL_FILE)
        logger.info(f"Создан новый файл {EXCEL_FILE}")
    
    return excel_path.exists()


def add_user_if_not_exists(user_id: str, username: str):
    """Добавляет пользователя в таблицу, если его там нет."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Проверяем, есть ли уже пользователь
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == user_id:
            wb.close()
            return
    
    # Добавляем нового пользователя
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value=user_id)
    ws.cell(row=new_row, column=2, value=0)
    
    wb.save(EXCEL_FILE)
    wb.close()
    logger.info(f"Добавлен пользователь: {username} ({user_id})")


def add_training_date_column(training_date: str):
    """Добавляет колонку для даты тренировки, если её нет."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Ищем колонку с этой датой
    for col in range(3, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == training_date:
            wb.close()
            return col
    
    # Добавляем новую колонку
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value=training_date)
    
    # Заполняем нулями существующих пользователей
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=new_col, value=0)
    
    wb.save(EXCEL_FILE)
    wb.close()
    logger.info(f"Добавлена колонка для даты: {training_date}")
    return new_col


def record_vote(user_id: str, username: str, training_date: str, vote_value: int):
    """Записывает результат голосования в таблицу."""
    add_user_if_not_exists(user_id, username)
    date_col = add_training_date_column(training_date)
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Находим строку пользователя
    user_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == user_id:
            user_row = row
            break
    
    if user_row:
        # Записываем результат голосования
        ws.cell(row=user_row, column=date_col, value=vote_value)
        
        # Пересчитываем посещаемость
        total_trainings = 0
        attended_trainings = 0
        for col in range(3, ws.max_column + 1):
            cell_value = ws.cell(row=user_row, column=col).value
            if cell_value is not None and isinstance(cell_value, (int, float)):
                total_trainings += 1
                if cell_value == 1:
                    attended_trainings += 1
        
        attendance_percent = round((attended_trainings / total_trainings * 100), 1) if total_trainings > 0 else 0
        ws.cell(row=user_row, column=2, value=attendance_percent)
        
        wb.save(EXCEL_FILE)
        logger.info(f"Записан голос пользователя {username}: {vote_value} на {training_date}")
    
    wb.close()


async def send_poll(application, title: str, training_date: str, time_str: str, location: str):
    """Отправляет опрос в группу."""
    try:
        poll = await application.bot.send_poll(
            chat_id=GROUP_ID,
            question=title,
            options=["иду", "плачу", "не иду"],
            is_anonymous=False,
            allows_multiple_answers=False,
            message_thread_id=THREAD_ID  # Отправка в конкретную ветку (если форум)
        )
        logger.info(f"Опрос отправлен: {title}")
        logger.info(f"ID опроса: {poll.poll.id}")
        
        # Сохраняем связь ID опроса и данных тренировки
        active_polls[poll.poll.id] = {
            "date": training_date,
            "time": time_str,
            "location": location
        }
        logger.info(f"Опрос {poll.poll.id} сохранен для даты {training_date} {time_str} {location}")
        
        return poll.poll.id
    except Exception as e:
        logger.error(f"Ошибка при отправке опроса: {e}")
        return None


async def send_friday_poll(application):
    """Отправляет опрос для пятничной тренировки."""
    friday_date = get_next_friday()
    date_str = friday_date.strftime('%d.%m')
    title = f"Тренировка {date_str} 18:00 БНТУ"
    await send_poll(application, title, date_str, "18:00", "БНТУ")


async def send_sunday_poll(application):
    """Отправляет опрос для воскресной тренировки."""
    sunday_date = get_next_sunday()
    date_str = sunday_date.strftime('%d.%m')
    title = f"Тренировка {date_str} 10:00 РГУОР"
    await send_poll(application, title, date_str, "10:00", "РГУОР")


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start."""
    await update.message.reply_text(
        "Привет! Я бот для записи на тренировки.\n"
        "Тренировки проходят по пятницам (18:00 БНТУ) и воскресеньям (10:00 РГУОР).\n"
        "Опросы отправляются за 4 дня до тренировки.\n"
        "Используйте команду /help для получения дополнительной информации."
    )


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /help."""
    await update.message.reply_text(
        "Доступные команды:\n"
        "/start - Начать работу с ботом\n"
        "/help - Показать эту справку\n"
        "/status - Показать текущую статистику посещаемости\n"
        "/test_poll - Тестовый опрос (только для администраторов)"
    )


async def status_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /status - показывает статистику."""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        if ws.max_row < 2:
            await update.message.reply_text("Пока нет данных о посещаемости.")
            wb.close()
            return
        
        # Сначала соберём все user_id из таблицы
        user_ids_in_table = []
        for row in range(2, ws.max_row + 1):
            user_id = ws.cell(row=row, column=1).value
            user_ids_in_table.append(str(user_id))
        
        message = "📊 Статистика посещаемости:\n\n"
        
        # Проходим по таблице и показываем данные
        for row in range(2, min(ws.max_row + 1, 11)):  # Показываем топ-10
            user_id = ws.cell(row=row, column=1).value
            attendance = ws.cell(row=row, column=2).value
            
            # Пытаемся получить имя пользователя, но не критично если не получится
            username = None
            try:
                user = await context.bot.get_chat(user_id)
                if user.first_name:
                    username = user.first_name
                    if user.last_name:
                        username += f" {user.last_name}"
                    if user.username:
                        username += f" (@{user.username})"
                elif user.username:
                    username = f"@{user.username}"
            except Exception as e:
                logger.debug(f"Не удалось получить имя для {user_id}: {e}")
            
            # Если не удалось получить имя, используем ID или сохраняем как есть
            if username is None:
                username = f"User_{user_id}"
            
            message += f"{username}: {attendance}%\n"
        
        if ws.max_row > 10:
            message += f"\n... и ещё {ws.max_row - 10} участников"
        
        await update.message.reply_text(message)
        wb.close()
    except FileNotFoundError:
        logger.error("Файл Excel не найден")
        await update.message.reply_text("Файл статистики не найден. Возможно, бот ещё не обрабатывал голосования.")
    except Exception as e:
        logger.error(f"Ошибка при получении статуса: {e}")
        await update.message.reply_text(f"Произошла ошибка при получении статистики: {str(e)}")


async def test_poll_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Тестовый опрос для проверки работы бота."""
    # Проверка прав администратора (упрощённая)
    chat_member = await context.bot.get_chat_member(
        chat_id=update.effective_chat.id,
        user_id=update.effective_user.id
    )
    
    if chat_member.status not in ['administrator', 'creator']:
        await update.message.reply_text("Эта команда доступна только администраторам.")
        return
    
    # Создаём тестовую дату
    test_date = datetime.now().strftime('%d.%m')
    title = f"ТЕСТ Тренировка {test_date}"
    
    await send_poll(context.application, title, test_date, "18:00", "БНТУ")
    await update.message.reply_text("Тестовый опрос отправлен!")


async def handle_poll_answer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик ответов на опросы."""
    poll_answer = update.poll_answer
    
    try:
        # В python-telegram-bot 20+ используем poll_answer.user.id вместо poll_answer.user_id
        user_id = poll_answer.user.id
        logger.info(f"Получен ответ на опрос от пользователя {user_id}")
        logger.info(f"ID опроса: {poll_answer.poll_id}")
        logger.info(f"Выбранные варианты: {poll_answer.option_ids}")
        
        # Получаем информацию об опросе из глобального хранилища или извлекаем из вопроса
        training_data = active_polls.get(poll_answer.poll_id)
        
        if training_data:
            date_str = training_data["date"]
            time_str = training_data["time"]
            location = training_data["location"]
            question = f"Тренировка {date_str} {time_str} {location}"
            logger.info(f"Найдены данные опроса: {question}")
        else:
            # Если опрос не найден в хранилище (например, после перезапуска бота),
            # пытаемся извлечь дату из текста вопроса через get_updates
            logger.warning(f"Опрос {poll_answer.poll_id} не найден в активном хранилище")
            date_str = datetime.now().strftime('%d.%m')
            logger.warning(f"Используем текущую дату: {date_str}")
        
        # Получаем информацию о пользователе
        user = await context.bot.get_chat(user_id)
        username = user.first_name or user.username or str(user_id)
        
        logger.info(f"Пользователь: {username} ({user_id})")
        
        # Определяем значение голоса
        if not poll_answer.option_ids:
            logger.warning("Пользователь не выбрал вариант ответа")
            return
            
        option_index = poll_answer.option_ids[0]
        # "иду" = 1, "плачу" = 0, "не иду" = 0
        vote_value = 1 if option_index == 0 else 0
        
        logger.info(f"Выбранный вариант: {option_index}, значение: {vote_value} ({'иду' if vote_value == 1 else 'не идёт'})")
        
        # Записываем в таблицу
        record_vote(str(user_id), username, date_str, vote_value)
        
        # Если пользователь идёт на тренировку, добавляем его в список attendees
        if vote_value == 1 and training_data:
            training_key = f"{date_str}_{time_str}"
            if training_key not in training_attendees:
                training_attendees[training_key] = {}
            training_attendees[training_key][str(user_id)] = username
            logger.info(f"Пользователь {username} записан на тренировку {training_key}")
        
        logger.info(f"✅ Голос успешно записан: {username} -> {'иду' if vote_value == 1 else 'не идёт'} на {date_str}")
        
    except Exception as e:
        logger.error(f"❌ Ошибка при обработке ответа на опрос: {e}", exc_info=True)


async def scheduled_task(context: ContextTypes.DEFAULT_TYPE):
    """Ежедневная проверка: нужно ли отправить опрос."""
    today = datetime.now().date()
    
    # Проверяем, какой сегодня день недели
    # Вторник (1) - отправляем опрос на пятницу (через 3 дня)
    # Среда (2) - отправляем опрос на воскресенье (через 4 дня)
    
    weekday = today.weekday()
    
    if weekday == 1:  # Вторник
        logger.info("Вторник - отправляем опрос на пятницу")
        await send_friday_poll(context.application)
    elif weekday == 2:  # Среда
        logger.info("Среда - отправляем опрос на воскресенье")
        await send_sunday_poll(context.application)


async def send_feedback_request(context: ContextTypes.DEFAULT_TYPE):
    """Отправляет запросы на фидбэк пользователям, которые были на тренировке."""
    now = datetime.now()
    current_hour = now.hour
    
    # Проверяем время для отправки фидбэка
    # Для тренировки в 18:00 - отправляем в 20:00
    # Для тренировки в 10:00 - отправляем в 12:00
    feedback_hours = {
        "18:00": 20,  # Пятница: тренировка в 18:00, фидбэк в 20:00
        "10:00": 12,  # Воскресенье: тренировка в 10:00, фидбэк в 12:00
    }
    
    # Определяем, для какого времени тренировки сейчас нужно отправить фидбэк
    target_time_str = None
    if current_hour == 20:
        target_time_str = "18:00"
    elif current_hour == 12:
        target_time_str = "10:00"
    
    if not target_time_str:
        return
    
    # Определяем дату тренировки (сегодня)
    training_date = now.strftime('%d.%m')
    training_key = f"{training_date}_{target_time_str}"
    
    # Проверяем, не отправляли ли уже фидбэк для этой тренировки
    if training_key in feedback_requests_sent:
        return
    
    # Проверяем, есть ли пользователи на этой тренировке
    if training_key not in training_attendees or not training_attendees[training_key]:
        return
    
    attendees = training_attendees[training_key]
    logger.info(f"Отправка фидбэка для тренировки {training_key}, участников: {len(attendees)}")
    
    for user_id, username in attendees.items():
        try:
            # Отправляем сообщение с просьбой оценить тренировку
            await context.bot.send_message(
                chat_id=user_id,
                text=f"Привет, {username}! 👋\n\n"
                     f"Тренировка закончилась. Пожалуйста, оцени её качество от 1 до 10.\n\n"
                     f"Просто напиши число (например: 7)"
            )
            logger.info(f"Запрос на оценку отправлен пользователю {username} ({user_id})")
        except Exception as e:
            logger.error(f"Ошибка при отправке запроса на оценку пользователю {user_id}: {e}")
    
    # Помечаем, что фидбэк для этой тренировки отправлен
    feedback_requests_sent.add(training_key)
    logger.info(f"Фидбэк для тренировки {training_key} отмечен как отправленный")


async def handle_feedback_rating(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик оценок тренировок от пользователей."""
    try:
        user_id = str(update.effective_user.id)
        message_text = update.message.text.strip()
        
        # Проверяем, является ли сообщение числом от 1 до 10
        try:
            rating = int(message_text)
            if rating < 1 or rating > 10:
                return  # Не является валидной оценкой
        except ValueError:
            return  # Не число
        
        # Сохраняем оценку
        save_feedback_rating(user_id, rating)
        
        # Отправляем подтверждение и просим написать фидбэк
        await update.message.reply_text(
            f"Спасибо за оценку: {rating}/10! ⭐\n\n"
            f"Если хочешь, можешь написать краткий отзыв о тренировке:\n"
            f"что понравилось, что можно улучшить и т.д.\n\n"
            f"(Если не хочешь писать отзыв, просто проигнорируй это сообщение)"
        )
        logger.info(f"Получена оценка {rating} от пользователя {user_id}")
        
    except Exception as e:
        logger.error(f"Ошибка при обработке оценки: {e}")


async def handle_feedback_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик текстового фидбэка от пользователей."""
    try:
        user_id = str(update.effective_user.id)
        message_text = update.message.text.strip()
        
        # Проверяем, не является ли это числом (оценкой)
        try:
            rating = int(message_text)
            if 1 <= rating <= 10:
                return  # Это оценка, а не текст фидбэка
        except ValueError:
            pass  # Это текст
        
        # Сохраняем текстовый фидбэк
        save_feedback_text(user_id, message_text)
        
        # Отправляем подтверждение
        await update.message.reply_text(
            "Спасибо за твой отзыв! 🙏\n"
            "Он поможет сделать тренировки лучше."
        )
        logger.info(f"Получен текстовый фидбэк от пользователя {user_id}: {message_text[:50]}...")
        
    except Exception as e:
        logger.error(f"Ошибка при обработке текстового фидбэка: {e}")


def init_feedback_excel():
    """Инициализирует Excel файл для хранения фидбэка."""
    excel_path = Path("feedback.xlsx")
    
    if not excel_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Фидбэк"
        
        # Заголовки
        ws.cell(row=1, column=1, value="Дата")
        ws.cell(row=1, column=2, value="Время")
        ws.cell(row=1, column=3, value="User ID")
        ws.cell(row=1, column=4, value="Имя")
        ws.cell(row=1, column=5, value="Оценка")
        ws.cell(row=1, column=6, value="Фидбэк")
        
        wb.save("feedback.xlsx")
        logger.info("Создан новый файл feedback.xlsx")
    
    return excel_path.exists()


def save_feedback_rating(user_id: str, rating: int):
    """Сохраняет оценку тренировки в Excel."""
    from openpyxl import load_workbook
    
    now = datetime.now()
    training_date = now.strftime('%d.%m')
    training_time = now.strftime('%H:%M')
    
    wb = load_workbook("feedback.xlsx")
    ws = wb.active
    
    # Получаем имя пользователя
    # (в реальном боте можно получить из контекста, здесь упрощённо)
    username = f"User_{user_id}"
    
    # Добавляем новую строку
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value=training_date)
    ws.cell(row=new_row, column=2, value=training_time)
    ws.cell(row=new_row, column=3, value=user_id)
    ws.cell(row=new_row, column=4, value=username)
    ws.cell(row=new_row, column=5, value=rating)
    ws.cell(row=new_row, column=6, value="")
    
    wb.save("feedback.xlsx")
    wb.close()
    logger.info(f"Сохранена оценка {rating} от пользователя {user_id}")


def save_feedback_text(user_id: str, feedback_text: str):
    """Сохраняет текстовый фидбэк в Excel."""
    from openpyxl import load_workbook
    
    now = datetime.now()
    training_date = now.strftime('%d.%m')
    training_time = now.strftime('%H:%M')
    
    wb = load_workbook("feedback.xlsx")
    ws = wb.active
    
    # Получаем имя пользователя
    username = f"User_{user_id}"
    
    # Ищем последнюю запись для этого пользователя за сегодня
    last_row_with_rating = None
    for row in range(ws.max_row, 1, -1):
        if (ws.cell(row=row, column=3).value == user_id and
            ws.cell(row=row, column=1).value == training_date):
            # Проверяем, есть ли уже оценка
            if ws.cell(row=row, column=5).value is not None:
                last_row_with_rating = row
                break
    
    if last_row_with_rating:
        # Обновляем существующую запись
        ws.cell(row=last_row_with_rating, column=6, value=feedback_text)
    else:
        # Добавляем новую запись (только текст, без оценки)
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=training_date)
        ws.cell(row=new_row, column=2, value=training_time)
        ws.cell(row=new_row, column=3, value=user_id)
        ws.cell(row=new_row, column=4, value=username)
        ws.cell(row=new_row, column=5, value="")
        ws.cell(row=new_row, column=6, value=feedback_text)
    
    wb.save("feedback.xlsx")
    wb.close()
    logger.info(f"Сохранён фидбэк от пользователя {user_id}: {feedback_text[:30]}...")


async def post_init(application):
    """Инициализация после запуска бота."""
    logger.info("Бот успешно запущен!")
    logger.info(f"Excel файл: {EXCEL_FILE}")
    logger.info(f"Группа: {GROUP_ID}")
    if THREAD_ID:
        logger.info(f"Ветка (topic): {THREAD_ID}")
    
    # Инициализация файла для фидбэка
    init_feedback_excel()
    logger.info("Файл feedback.xlsx инициализирован")


def main():
    """Основная функция запуска бота."""
    # Инициализация Excel файла
    init_excel_file()
    
    # Создание приложения
    application = Application.builder().token(BOT_TOKEN).post_init(post_init).build()
    
    # Добавление обработчиков команд
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("status", status_command))
    application.add_handler(CommandHandler("test_poll", test_poll_command))
    
    # Обработчик ответов на опросы
    application.add_handler(PollAnswerHandler(handle_poll_answer))
    
    # Обработчики текстовых сообщений для фидбэка
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_feedback_text))
    
    # Добавляем ежедневную задачу в 10:00 для отправки опросов
    job_queue = application.job_queue
    job_queue.run_daily(scheduled_task, time=datetime.strptime("10:00", "%H:%M").time())
    logger.info("Запланирована ежедневная задача на 10:00 (опросы)")
    
    # Добавляем задачи для отправки фидбэка
    # В 12:00 - для воскресной тренировки (10:00)
    job_queue.run_daily(send_feedback_request, time=datetime.strptime("12:00", "%H:%M").time())
    logger.info("Запланирована ежедневная задача на 12:00 (фидбэк для тренировки 10:00)")
    
    # В 20:00 - для пятничной тренировки (18:00)
    job_queue.run_daily(send_feedback_request, time=datetime.strptime("20:00", "%H:%M").time())
    logger.info("Запланирована ежедневная задача на 20:00 (фидбэк для тренировки 18:00)")
    
    # Запуск бота
    logger.info("Бот запущен и ожидает команды...")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
